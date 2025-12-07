"""
Utilitaires d'export pour l'application de comptabilité
Export vers Excel, PDF, CSV
"""
import os
from datetime import datetime
from typing import List, Dict, Any, Optional
from decimal import Decimal
import logging

logger = logging.getLogger(__name__)

# Imports optionnels (avec fallback si non installés)
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("⚠️ openpyxl non installé - Export Excel désactivé")

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logger.warning("⚠️ reportlab non installé - Export PDF désactivé")


class ExportManager:
    """Gestionnaire d'export de données comptables"""

    def __init__(self, output_dir: str = "/tmp"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def _format_montant(self, montant: Any) -> str:
        """Formate un montant pour l'affichage"""
        if isinstance(montant, (Decimal, float)):
            return f"{float(montant):,.2f}".replace(',', ' ')
        return str(montant)

    def _format_date(self, date_obj: Any) -> str:
        """Formate une date pour l'affichage"""
        if date_obj:
            if hasattr(date_obj, 'strftime'):
                return date_obj.strftime("%d/%m/%Y")
        return str(date_obj) if date_obj else ""

    # ========== EXPORT EXCEL ==========

    def exporter_balance_excel(
        self,
        balance_data: List[Dict],
        societe_nom: str,
        exercice_annee: int,
        filename: Optional[str] = None
    ) -> tuple[bool, str]:
        """
        Exporte la balance vers Excel
        Returns: (succès, chemin_fichier_ou_message_erreur)
        """
        if not EXCEL_AVAILABLE:
            return False, "❌ openpyxl n'est pas installé. Installez-le avec: pip install openpyxl"

        try:
            if not filename:
                filename = f"Balance_{societe_nom}_{exercice_annee}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            filepath = os.path.join(self.output_dir, filename)

            # Créer le classeur
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Balance"

            # Styles
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            title_font = Font(bold=True, size=14)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Titre
            ws.merge_cells('A1:G1')
            cell = ws['A1']
            cell.value = f"BALANCE GÉNÉRALE - {societe_nom} - Exercice {exercice_annee}"
            cell.font = title_font
            cell.alignment = Alignment(horizontal="center")

            # Date d'édition
            ws.merge_cells('A2:G2')
            cell = ws['A2']
            cell.value = f"Édité le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
            cell.alignment = Alignment(horizontal="center")

            # En-têtes
            headers = ["Compte", "Intitulé", "Classe", "Débit", "Crédit", "Solde Débiteur", "Solde Créditeur"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Données
            total_debit = Decimal('0')
            total_credit = Decimal('0')
            total_solde_debiteur = Decimal('0')
            total_solde_crediteur = Decimal('0')

            row = 5
            for item in balance_data:
                solde = Decimal(str(item.get('solde', 0)))
                debit = Decimal(str(item.get('total_debit', 0)))
                credit = Decimal(str(item.get('total_credit', 0)))

                ws.cell(row=row, column=1, value=item.get('compte', ''))
                ws.cell(row=row, column=2, value=item.get('intitule', ''))
                ws.cell(row=row, column=3, value=item.get('classe', ''))
                ws.cell(row=row, column=4, value=float(debit))
                ws.cell(row=row, column=5, value=float(credit))

                if solde > 0:
                    ws.cell(row=row, column=6, value=float(solde))
                    total_solde_debiteur += solde
                else:
                    ws.cell(row=row, column=7, value=float(abs(solde)))
                    total_solde_crediteur += abs(solde)

                total_debit += debit
                total_credit += credit

                # Bordures
                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = border

                row += 1

            # Ligne de totaux
            row += 1
            ws.cell(row=row, column=1, value="TOTAUX").font = Font(bold=True)
            ws.cell(row=row, column=4, value=float(total_debit)).font = Font(bold=True)
            ws.cell(row=row, column=5, value=float(total_credit)).font = Font(bold=True)
            ws.cell(row=row, column=6, value=float(total_solde_debiteur)).font = Font(bold=True)
            ws.cell(row=row, column=7, value=float(total_solde_crediteur)).font = Font(bold=True)

            for col in range(1, 8):
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

            # Format des nombres
            for row_idx in range(5, row + 1):
                for col in [4, 5, 6, 7]:
                    ws.cell(row=row_idx, column=col).number_format = '#,##0.00'

            # Ajuster la largeur des colonnes
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 18

            # Sauvegarder
            wb.save(filepath)
            logger.info(f"✅ Balance exportée vers Excel : {filepath}")
            return True, filepath

        except Exception as e:
            logger.error(f"❌ Erreur export Excel : {e}", exc_info=True)
            return False, f"❌ Erreur : {str(e)}"

    def exporter_compte_resultat_excel(
        self,
        charges: List[Dict],
        produits: List[Dict],
        total_charges: Decimal,
        total_produits: Decimal,
        resultat: Decimal,
        societe_nom: str,
        exercice_annee: int,
        filename: Optional[str] = None
    ) -> tuple[bool, str]:
        """Exporte le compte de résultat vers Excel"""
        if not EXCEL_AVAILABLE:
            return False, "❌ openpyxl n'est pas installé"

        try:
            if not filename:
                filename = f"CompteResultat_{societe_nom}_{exercice_annee}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            filepath = os.path.join(self.output_dir, filename)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Compte de Résultat"

            # Titre
            ws.merge_cells('A1:D1')
            cell = ws['A1']
            cell.value = f"COMPTE DE RÉSULTAT - {societe_nom} - Exercice {exercice_annee}"
            cell.font = Font(bold=True, size=14)
            cell.alignment = Alignment(horizontal="center")

            row = 3

            # CHARGES
            ws.cell(row=row, column=1, value="CHARGES").font = Font(bold=True, size=12)
            row += 1

            ws.cell(row=row, column=1, value="Compte").font = Font(bold=True)
            ws.cell(row=row, column=2, value="Intitulé").font = Font(bold=True)
            ws.cell(row=row, column=3, value="Montant").font = Font(bold=True)
            row += 1

            for charge in charges:
                ws.cell(row=row, column=1, value=charge.get('compte', ''))
                ws.cell(row=row, column=2, value=charge.get('intitule', ''))
                ws.cell(row=row, column=3, value=float(charge.get('solde', 0)))
                ws.cell(row=row, column=3).number_format = '#,##0.00'
                row += 1

            ws.cell(row=row, column=2, value="Total Charges").font = Font(bold=True)
            ws.cell(row=row, column=3, value=float(total_charges)).font = Font(bold=True)
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            row += 2

            # PRODUITS
            ws.cell(row=row, column=1, value="PRODUITS").font = Font(bold=True, size=12)
            row += 1

            ws.cell(row=row, column=1, value="Compte").font = Font(bold=True)
            ws.cell(row=row, column=2, value="Intitulé").font = Font(bold=True)
            ws.cell(row=row, column=3, value="Montant").font = Font(bold=True)
            row += 1

            for produit in produits:
                ws.cell(row=row, column=1, value=produit.get('compte', ''))
                ws.cell(row=row, column=2, value=produit.get('intitule', ''))
                ws.cell(row=row, column=3, value=float(produit.get('solde', 0)))
                ws.cell(row=row, column=3).number_format = '#,##0.00'
                row += 1

            ws.cell(row=row, column=2, value="Total Produits").font = Font(bold=True)
            ws.cell(row=row, column=3, value=float(total_produits)).font = Font(bold=True)
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            row += 2

            # RÉSULTAT
            ws.cell(row=row, column=2, value="RÉSULTAT DE L'EXERCICE").font = Font(bold=True, size=12)
            ws.cell(row=row, column=3, value=float(resultat)).font = Font(bold=True, size=12)
            ws.cell(row=row, column=3).number_format = '#,##0.00'

            if resultat >= 0:
                ws.cell(row=row, column=3).font = Font(bold=True, color="008000", size=12)
            else:
                ws.cell(row=row, column=3).font = Font(bold=True, color="FF0000", size=12)

            # Largeurs des colonnes
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 18

            wb.save(filepath)
            logger.info(f"✅ Compte de résultat exporté : {filepath}")
            return True, filepath

        except Exception as e:
            logger.error(f"❌ Erreur export Excel : {e}", exc_info=True)
            return False, f"❌ Erreur : {str(e)}"

    # ========== EXPORT CSV ==========

    def exporter_balance_csv(
        self,
        balance_data: List[Dict],
        filename: Optional[str] = None
    ) -> tuple[bool, str]:
        """Exporte la balance vers CSV"""
        try:
            import csv

            if not filename:
                filename = f"balance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

            filepath = os.path.join(self.output_dir, filename)

            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f, delimiter=';')

                # En-tête
                writer.writerow(['Compte', 'Intitulé', 'Classe', 'Débit', 'Crédit', 'Solde'])

                # Données
                for item in balance_data:
                    writer.writerow([
                        item.get('compte', ''),
                        item.get('intitule', ''),
                        item.get('classe', ''),
                        self._format_montant(item.get('total_debit', 0)),
                        self._format_montant(item.get('total_credit', 0)),
                        self._format_montant(item.get('solde', 0))
                    ])

            logger.info(f"✅ Balance exportée vers CSV : {filepath}")
            return True, filepath

        except Exception as e:
            logger.error(f"❌ Erreur export CSV : {e}", exc_info=True)
            return False, f"❌ Erreur : {str(e)}"


def installer_dependances_export():
    """
    Affiche les commandes pour installer les dépendances d'export
    """
    print("\n" + "="*60)
    print("DÉPENDANCES POUR L'EXPORT")
    print("="*60)

    if not EXCEL_AVAILABLE:
        print("\n📊 Pour l'export Excel:")
        print("   pip install openpyxl")

    if not PDF_AVAILABLE:
        print("\n📄 Pour l'export PDF:")
        print("   pip install reportlab")

    if EXCEL_AVAILABLE and PDF_AVAILABLE:
        print("\n✅ Toutes les dépendances d'export sont installées!")

    print("="*60 + "\n")
