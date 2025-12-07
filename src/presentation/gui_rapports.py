"""
Fenêtres de rapports comptables - VERSION COMPLÈTE ET CORRIGÉE
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from decimal import Decimal
from datetime import datetime


class BalanceWindow:
    """Fenêtre d'affichage de la balance"""
    
    def __init__(self, parent, service, societe, exercice):
        self.service = service
        self.societe = societe
        self.exercice = exercice
        
        self.window = tk.Toplevel(parent)
        self.window.title(f"Balance - {exercice.annee}")
        self.window.geometry("1000x600")
        
        self.create_widgets()
        self.load_balance()
    
    def create_widgets(self):
        """Crée les widgets"""
        main_frame = ttk.Frame(self.window, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Titre
        ttk.Label(main_frame, text=f"⚖️ BALANCE - {self.exercice.annee}", 
                 font=('Arial', 14, 'bold')).grid(row=0, column=0, pady=(0, 10))
        
        # TreeView
        columns = ('compte', 'intitule', 'debit', 'credit', 'solde')
        self.tree = ttk.Treeview(main_frame, columns=columns, show='headings', height=25)
        
        self.tree.heading('compte', text='Compte')
        self.tree.heading('intitule', text='Intitulé')
        self.tree.heading('debit', text='Débit')
        self.tree.heading('credit', text='Crédit')
        self.tree.heading('solde', text='Solde')
        
        self.tree.column('compte', width=100)
        self.tree.column('intitule', width=350)
        self.tree.column('debit', width=120, anchor=tk.E)
        self.tree.column('credit', width=120, anchor=tk.E)
        self.tree.column('solde', width=120, anchor=tk.E)
        
        self.tree.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar.grid(row=1, column=1, sticky=(tk.N, tk.S))
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        # Totaux
        totaux_frame = ttk.Frame(main_frame)
        totaux_frame.grid(row=2, column=0, pady=10, sticky=(tk.W, tk.E))
        
        self.lbl_totaux = ttk.Label(totaux_frame, text="", font=('Arial', 10, 'bold'))
        self.lbl_totaux.pack()

        # Boutons d'action
        btn_frame = ttk.Frame(main_frame)
        btn_frame.grid(row=3, column=0, pady=10)

        ttk.Button(btn_frame, text="📊 Export Excel", command=self.export_excel).grid(row=0, column=0, padx=5)
        ttk.Button(btn_frame, text="📄 Export CSV", command=self.export_csv).grid(row=0, column=1, padx=5)
        ttk.Button(btn_frame, text="Fermer", command=self.window.destroy).grid(row=0, column=2, padx=5)
        
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
        self.window.columnconfigure(0, weight=1)
        self.window.rowconfigure(0, weight=1)
    
    def load_balance(self):
        """Charge la balance"""
        balance = self.service.get_balance(self.societe.id, self.exercice.id)
        
        total_debit = Decimal('0')
        total_credit = Decimal('0')
        
        for ligne in balance:
            self.tree.insert('', tk.END, values=(
                ligne.compte,
                ligne.intitule,
                f"{ligne.total_debit:.2f}",
                f"{ligne.total_credit:.2f}",
                f"{ligne.solde:.2f}"
            ))
            total_debit += ligne.total_debit
            total_credit += ligne.total_credit
        
        self.lbl_totaux.config(
            text=f"TOTAUX : Débit = {total_debit:.2f} € | Crédit = {total_credit:.2f} €"
        )

    def export_excel(self):
        """Exporte la balance en Excel"""
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Tous les fichiers", "*.*")],
            initialfile=f"balance_{self.exercice.annee}.xlsx"
        )

        if not filename:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Balance"

            # En-tête
            ws['A1'] = f"BALANCE - {self.societe.nom}"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A2'] = f"Exercice {self.exercice.annee}"
            ws['A2'].font = Font(size=12)
            ws['A3'] = f"Édité le {datetime.now().strftime('%d/%m/%Y %H:%M')}"

            # Colonnes
            row = 5
            headers = ['Compte', 'Intitulé', 'Débit', 'Crédit', 'Solde']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Données
            total_debit = Decimal('0')
            total_credit = Decimal('0')

            for item in self.tree.get_children():
                row += 1
                values = self.tree.item(item)['values']

                ws.cell(row=row, column=1, value=values[0])
                ws.cell(row=row, column=2, value=values[1])
                ws.cell(row=row, column=3, value=float(values[2]))
                ws.cell(row=row, column=4, value=float(values[3]))
                ws.cell(row=row, column=5, value=float(values[4]))

                total_debit += Decimal(str(values[2]))
                total_credit += Decimal(str(values[3]))

            # Totaux
            row += 1
            ws.cell(row=row, column=2, value="TOTAUX").font = Font(bold=True)
            ws.cell(row=row, column=3, value=float(total_debit)).font = Font(bold=True)
            ws.cell(row=row, column=4, value=float(total_credit)).font = Font(bold=True)

            # Ajuster largeurs
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15

            wb.save(filename)
            messagebox.showinfo("Succès", f"✅ Export Excel réussi\n{filename}")

        except ImportError:
            messagebox.showerror("Erreur", "La bibliothèque openpyxl n'est pas installée")
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'export: {e}")

    def export_csv(self):
        """Exporte la balance en CSV"""
        filename = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv"), ("Tous les fichiers", "*.*")],
            initialfile=f"balance_{self.exercice.annee}.csv"
        )

        if not filename:
            return

        try:
            with open(filename, 'w', encoding='utf-8') as f:
                # En-tête
                f.write(f"BALANCE - {self.societe.nom}\n")
                f.write(f"Exercice {self.exercice.annee}\n")
                f.write(f"Édité le {datetime.now().strftime('%d/%m/%Y %H:%M')}\n")
                f.write("\n")

                # Colonnes
                f.write("Compte;Intitulé;Débit;Crédit;Solde\n")

                # Données
                for item in self.tree.get_children():
                    values = self.tree.item(item)['values']
                    f.write(";".join(str(v) for v in values) + "\n")

            messagebox.showinfo("Succès", f"✅ Export CSV réussi\n{filename}")

        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'export: {e}")


class ResultatWindow:
    """Fenêtre du compte de résultat"""
    
    def __init__(self, parent, service, societe, exercice):
        self.service = service
        self.societe = societe
        self.exercice = exercice
        
        self.window = tk.Toplevel(parent)
        self.window.title(f"Compte de Résultat - {exercice.annee}")
        self.window.geometry("900x700")
        
        self.create_widgets()
        self.load_resultat()
    
    def create_widgets(self):
        """Crée les widgets"""
        main_frame = ttk.Frame(self.window, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Titre
        ttk.Label(main_frame, text=f"📊 COMPTE DE RÉSULTAT - {self.exercice.annee}", 
                 font=('Arial', 14, 'bold')).grid(row=0, column=0, columnspan=2, pady=(0, 10))
        
        # Frame charges
        charges_frame = ttk.LabelFrame(main_frame, text="CHARGES (Classe 6)", padding="10")
        charges_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 5))
        
        columns = ('compte', 'intitule', 'montant')
        self.tree_charges = ttk.Treeview(charges_frame, columns=columns, show='headings', height=20)
        
        self.tree_charges.heading('compte', text='Compte')
        self.tree_charges.heading('intitule', text='Intitulé')
        self.tree_charges.heading('montant', text='Montant')
        
        self.tree_charges.column('compte', width=80)
        self.tree_charges.column('intitule', width=250)
        self.tree_charges.column('montant', width=100, anchor=tk.E)
        
        self.tree_charges.pack(fill=tk.BOTH, expand=True)
        
        # Frame produits
        produits_frame = ttk.LabelFrame(main_frame, text="PRODUITS (Classe 7)", padding="10")
        produits_frame.grid(row=1, column=1, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(5, 0))
        
        self.tree_produits = ttk.Treeview(produits_frame, columns=columns, show='headings', height=20)
        
        self.tree_produits.heading('compte', text='Compte')
        self.tree_produits.heading('intitule', text='Intitulé')
        self.tree_produits.heading('montant', text='Montant')
        
        self.tree_produits.column('compte', width=80)
        self.tree_produits.column('intitule', width=250)
        self.tree_produits.column('montant', width=100, anchor=tk.E)
        
        self.tree_produits.pack(fill=tk.BOTH, expand=True)
        
        # Résumé
        resume_frame = ttk.Frame(main_frame)
        resume_frame.grid(row=2, column=0, columnspan=2, pady=20)
        
        self.lbl_charges = ttk.Label(resume_frame, text="Total charges: 0.00 €", 
                                     font=('Arial', 11, 'bold'))
        self.lbl_charges.grid(row=0, column=0, padx=20)
        
        self.lbl_produits = ttk.Label(resume_frame, text="Total produits: 0.00 €", 
                                      font=('Arial', 11, 'bold'))
        self.lbl_produits.grid(row=0, column=1, padx=20)
        
        self.lbl_resultat = ttk.Label(resume_frame, text="Résultat: 0.00 €", 
                                      font=('Arial', 14, 'bold'))
        self.lbl_resultat.grid(row=1, column=0, columnspan=2, pady=10)
        
        # Bouton
        ttk.Button(main_frame, text="Fermer", command=self.window.destroy).grid(
            row=3, column=0, columnspan=2, pady=10)
        
        main_frame.columnconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
        self.window.columnconfigure(0, weight=1)
        self.window.rowconfigure(0, weight=1)
    
    def load_resultat(self):
        """Charge le compte de résultat"""
        data = self.service.get_compte_resultat(self.societe.id, self.exercice.id)
        
        # Charges
        for ligne in data['charges']:
            self.tree_charges.insert('', tk.END, values=(
                ligne['compte'],
                ligne['intitule'],
                f"{abs(float(ligne['solde'])):.2f}"
            ))
        
        # Produits
        for ligne in data['produits']:
            self.tree_produits.insert('', tk.END, values=(
                ligne['compte'],
                ligne['intitule'],
                f"{abs(float(ligne['solde'])):.2f}"
            ))
        
        # Totaux
        self.lbl_charges.config(text=f"Total charges: {abs(data['total_charges']):.2f} €")
        self.lbl_produits.config(text=f"Total produits: {abs(data['total_produits']):.2f} €")
        
        resultat = data['resultat']
        couleur = 'green' if resultat > 0 else 'red'
        type_resultat = "BÉNÉFICE" if resultat > 0 else "PERTE"
        
        self.lbl_resultat.config(
            text=f"Résultat ({type_resultat}): {abs(resultat):.2f} €",
            foreground=couleur
        )


class BilanWindow:
    """
    Fenêtre du bilan - VERSION CORRIGÉE
    
    Classification basée sur le TYPE de compte (actif/passif) et non sur le solde
    """
    
    def __init__(self, parent, service, societe, exercice):
        self.service = service
        self.societe = societe
        self.exercice = exercice
        
        self.window = tk.Toplevel(parent)
        self.window.title(f"Bilan - {exercice.annee}")
        self.window.geometry("1000x700")
        
        self.create_widgets()
        self.load_bilan()
    
    def create_widgets(self):
        """Crée les widgets"""
        main_frame = ttk.Frame(self.window, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Titre
        title_frame = ttk.Frame(main_frame)
        title_frame.grid(row=0, column=0, columnspan=2, pady=(0, 10))
        
        ttk.Label(title_frame, text=f"📋 BILAN - {self.exercice.annee}", 
                 font=('Arial', 14, 'bold')).pack()
        ttk.Label(title_frame, text="Classification par TYPE de compte (actif/passif)", 
                 font=('Arial', 9, 'italic'), foreground='blue').pack()
        
        # Frame actif
        actif_frame = ttk.LabelFrame(main_frame, text="ACTIF (Classes 2, 3, 4-actif, 5-actif)", padding="10")
        actif_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 5))
        
        columns = ('compte', 'intitule', 'type', 'montant')
        self.tree_actif = ttk.Treeview(actif_frame, columns=columns, show='headings', height=20)
        
        self.tree_actif.heading('compte', text='Compte')
        self.tree_actif.heading('intitule', text='Intitulé')
        self.tree_actif.heading('type', text='Solde')
        self.tree_actif.heading('montant', text='Montant')
        
        self.tree_actif.column('compte', width=80)
        self.tree_actif.column('intitule', width=250)
        self.tree_actif.column('type', width=60)
        self.tree_actif.column('montant', width=100, anchor=tk.E)
        
        # Scrollbar actif
        scrollbar_actif = ttk.Scrollbar(actif_frame, orient=tk.VERTICAL, command=self.tree_actif.yview)
        self.tree_actif.configure(yscrollcommand=scrollbar_actif.set)
        
        self.tree_actif.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_actif.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Frame passif
        passif_frame = ttk.LabelFrame(main_frame, text="PASSIF (Classe 1, 4-passif, 5-passif)", padding="10")
        passif_frame.grid(row=1, column=1, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(5, 0))
        
        self.tree_passif = ttk.Treeview(passif_frame, columns=columns, show='headings', height=20)
        
        self.tree_passif.heading('compte', text='Compte')
        self.tree_passif.heading('intitule', text='Intitulé')
        self.tree_passif.heading('type', text='Solde')
        self.tree_passif.heading('montant', text='Montant')
        
        self.tree_passif.column('compte', width=80)
        self.tree_passif.column('intitule', width=250)
        self.tree_passif.column('type', width=60)
        self.tree_passif.column('montant', width=100, anchor=tk.E)
        
        # Scrollbar passif
        scrollbar_passif = ttk.Scrollbar(passif_frame, orient=tk.VERTICAL, command=self.tree_passif.yview)
        self.tree_passif.configure(yscrollcommand=scrollbar_passif.set)
        
        self.tree_passif.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_passif.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Totaux
        totaux_frame = ttk.Frame(main_frame)
        totaux_frame.grid(row=2, column=0, columnspan=2, pady=20)
        
        self.lbl_actif = ttk.Label(totaux_frame, text="Total actif: 0.00 €", 
                                   font=('Arial', 12, 'bold'))
        self.lbl_actif.grid(row=0, column=0, padx=30)
        
        self.lbl_passif = ttk.Label(totaux_frame, text="Total passif: 0.00 €", 
                                    font=('Arial', 12, 'bold'))
        self.lbl_passif.grid(row=0, column=1, padx=30)
        
        self.lbl_equilibre = ttk.Label(totaux_frame, text="", 
                                       font=('Arial', 10))
        self.lbl_equilibre.grid(row=1, column=0, columnspan=2, pady=5)
        
        # Bouton
        ttk.Button(main_frame, text="Fermer", command=self.window.destroy).grid(
            row=3, column=0, columnspan=2, pady=10)
        
        main_frame.columnconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
        self.window.columnconfigure(0, weight=1)
        self.window.rowconfigure(0, weight=1)
    
    def load_bilan(self):
        """
        Charge le bilan - VERSION CORRIGÉE
        Classification basée sur le TYPE de compte et non sur le solde
        """
        try:
            data = self.service.get_bilan(self.societe.id, self.exercice.id)
            
            total_actif = Decimal('0')
            total_passif = Decimal('0')
            
            # ACTIF - Comptes de TYPE 'actif'
            for ligne in data['actif']:
                montant = abs(float(ligne['solde']))
                total_actif += Decimal(str(montant))
                
                # Indicateur du type de solde (pour information)
                solde_brut = float(ligne['solde'])
                type_solde = "Déb." if solde_brut >= 0 else "Créd."
                
                # Mettre en évidence les soldes inhabituels
                tags = ()
                if solde_brut < 0:  # Solde créditeur sur un compte actif = inhabituel
                    tags = ('warning',)
                
                self.tree_actif.insert('', tk.END, values=(
                    ligne['compte'],
                    ligne['intitule'],
                    type_solde,
                    f"{montant:.2f}"
                ), tags=tags)
            
            # PASSIF - Comptes de TYPE 'passif'
            for ligne in data['passif']:
                montant = abs(float(ligne['solde']))
                total_passif += Decimal(str(montant))
                
                # Indicateur du type de solde (pour information)
                solde_brut = float(ligne['solde'])
                type_solde = "Créd." if solde_brut <= 0 else "Déb."
                
                # Mettre en évidence les soldes inhabituels
                tags = ()
                if solde_brut > 0:  # Solde débiteur sur un compte passif = inhabituel
                    tags = ('warning',)
                
                self.tree_passif.insert('', tk.END, values=(
                    ligne['compte'],
                    ligne['intitule'],
                    type_solde,
                    f"{montant:.2f}"
                ), tags=tags)
            
            # Configuration des tags pour les soldes inhabituels
            self.tree_actif.tag_configure('warning', background='#fff3cd')
            self.tree_passif.tag_configure('warning', background='#fff3cd')
            
            # Totaux
            self.lbl_actif.config(text=f"Total actif: {total_actif:.2f} €")
            self.lbl_passif.config(text=f"Total passif: {total_passif:.2f} €")
            
            # Vérification de l'équilibre
            difference = abs(total_actif - total_passif)
            if difference < 0.01:  # Tolérance pour les arrondis
                self.lbl_equilibre.config(
                    text="✅ Bilan équilibré",
                    foreground='green'
                )
            else:
                self.lbl_equilibre.config(
                    text=f"⚠️ Déséquilibre: {difference:.2f} € (vérifier le résultat comptabilisé)",
                    foreground='orange'
                )
                
        except Exception as e:
            messagebox.showerror(
                "Erreur",
                f"Impossible de charger le bilan:\n{str(e)}\n\n"
                "Vérifiez que la méthode get_bilan() dans services.py "
                "classe bien les comptes par TYPE et non par solde."
            )


class TVAWindow:
    """Fenêtre récapitulatif TVA"""
    
    def __init__(self, parent, service, societe, exercice):
        self.service = service
        self.societe = societe
        self.exercice = exercice
        
        self.window = tk.Toplevel(parent)
        self.window.title(f"TVA - {exercice.annee}")
        self.window.geometry("700x500")
        
        self.create_widgets()
        self.load_tva()
    
    def create_widgets(self):
        """Crée les widgets"""
        main_frame = ttk.Frame(self.window, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Titre
        ttk.Label(main_frame, text=f"💶 RÉCAPITULATIF TVA - {self.exercice.annee}", 
                 font=('Arial', 14, 'bold')).grid(row=0, column=0, pady=(0, 20))
        
        # TreeView
        columns = ('type', 'compte', 'intitule', 'montant')
        self.tree = ttk.Treeview(main_frame, columns=columns, show='headings', height=15)
        
        self.tree.heading('type', text='Type')
        self.tree.heading('compte', text='Compte')
        self.tree.heading('intitule', text='Intitulé')
        self.tree.heading('montant', text='Montant')
        
        self.tree.column('type', width=150)
        self.tree.column('compte', width=100)
        self.tree.column('intitule', width=250)
        self.tree.column('montant', width=120, anchor=tk.E)
        
        self.tree.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Résumé
        resume_frame = ttk.LabelFrame(main_frame, text="Résumé", padding="15")
        resume_frame.grid(row=2, column=0, pady=20, sticky=(tk.W, tk.E))
        
        self.lbl_collectee = ttk.Label(resume_frame, text="TVA Collectée: 0.00 €", 
                                       font=('Arial', 11))
        self.lbl_collectee.grid(row=0, column=0, sticky=tk.W, pady=5)
        
        self.lbl_deductible = ttk.Label(resume_frame, text="TVA Déductible: 0.00 €", 
                                        font=('Arial', 11))
        self.lbl_deductible.grid(row=1, column=0, sticky=tk.W, pady=5)
        
        ttk.Separator(resume_frame, orient='horizontal').grid(
            row=2, column=0, sticky=(tk.W, tk.E), pady=10)
        
        self.lbl_a_payer = ttk.Label(resume_frame, text="TVA À PAYER: 0.00 €", 
                                     font=('Arial', 13, 'bold'), foreground='blue')
        self.lbl_a_payer.grid(row=3, column=0, sticky=tk.W, pady=5)
        
        # Bouton
        ttk.Button(main_frame, text="Fermer", command=self.window.destroy).grid(
            row=3, column=0, pady=10)
        
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
        self.window.columnconfigure(0, weight=1)
        self.window.rowconfigure(0, weight=1)
    
    def load_tva(self):
        """Charge le récapitulatif TVA"""
        data = self.service.get_tva_recap(self.societe.id, self.exercice.id)
        
        # Lignes de détail
        for ligne in data['lignes']:
            self.tree.insert('', tk.END, values=(
                ligne['type_tva'] or '',
                ligne['compte'],
                ligne['intitule'],
                f"{abs(float(ligne['solde'])):.2f}"
            ))
        
        # Résumé
        self.lbl_collectee.config(text=f"TVA Collectée: {data['tva_collectee']:.2f} €")
        self.lbl_deductible.config(text=f"TVA Déductible: {data['tva_deductible']:.2f} €")
        self.lbl_a_payer.config(text=f"TVA À PAYER: {data['tva_a_payer']:.2f} €")